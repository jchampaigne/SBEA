#!/usr/bin/env python
# coding: utf-8

# # SBEA Website Statistics Program

# In[1]:


import pandas as pd
import os
import numpy as np
import tkinter as tk
from tkinter import filedialog, messagebox
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter
from datetime import datetime
from openpyxl.styles import PatternFill


# In[2]:


# ---- UTILITY ----

def select_file(title):
    file_path = filedialog.askopenfilename(
        title=title,
        filetypes=[("CSV files", "*.csv")]
    )
    return file_path

def convert_file_to_UTF(file_path):
    try:
        with open(file_path, "r", encoding="windows-1252") as infile:
            content = infile.read()
        with open(file_path, "w", encoding="utf-8") as outfile:
            outfile.write(content)
        print(f"Converted {file_path} to UTF-8")
    except UnicodeDecodeError:
        print(f"Skipping {file_path} due to decoding error")

def save_to_excel(df, default_name="SBEA_statistics.xlsx"):
    file_path = filedialog.asksaveasfilename(
        title="Save SBEA Statistics Excel File",
        defaultextension=".xlsx",
        initialfile=default_name,
        filetypes=[("Excel files", "*.xlsx")]
    )
    if not file_path:
        print("Save cancelled.")
        return

    # Save to Excel
    df.to_excel(file_path, index=False)

    # Load workbook
    wb = load_workbook(file_path)
    ws = wb.active

    max_row = ws.max_row
    max_col = ws.max_column
    total_row = max_row

    # Format headers
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(wrap_text=True, horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 42

    # Bold column A (school names)
    for row in range(2, max_row + 1):
        ws[f"A{row}"].font = Font(bold=True)

    # Set font size
    for row in range(1, ws.max_row + 1):
        for col in range(1, ws.max_column + 1):
            ws.cell(row=row, column=col).font = Font(size=16)
    
    # Format numbers
    for col in range(2, max_col + 1):
        col_letter = get_column_letter(col)
        for row in range(2, max_row + 1):
            ws[f"{col_letter}{row}"].number_format = '#,##0'

    # Set column width
    for col in range(1, ws.max_column + 1):
        col_letter = get_column_letter(col)
        ws.column_dimensions[col_letter].width = 20

    # Full thin borders for the main table (including all edges)
    thin_side = Side(style='thin')
    thin_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

    for row in range(1, max_row + 1):
        for col in range(1, max_col + 1):
            cell = ws.cell(row=row, column=col)
            cell.border = thin_border

    # Double line above TOTAL row (last data row)
    double_top = Side(style="double")
    for col in range(1, max_col + 1):
        cell = ws.cell(row=total_row, column=col)
        cell.border = Border(
            top=double_top,
            left=cell.border.left,
            right=cell.border.right,
            bottom=cell.border.bottom
        )

    # Thick borders
    thick = Side(style='thick')

    # Around entire table
    for col in range(1, max_col + 1):
        # Top row
        top_cell = ws.cell(row=1, column=col)
        top_cell.border = Border(
            top=thick, left=top_cell.border.left,
            right=top_cell.border.right, bottom=top_cell.border.bottom
        )
        # Bottom row
        bottom_cell = ws.cell(row=max_row, column=col)
        bottom_cell.border = Border(
            bottom=thick, left=bottom_cell.border.left,
            right=bottom_cell.border.right, top=bottom_cell.border.top
        )
    for row in range(1, max_row + 1):
        # Left column
        left_cell = ws.cell(row=row, column=1)
        left_cell.border = Border(
            left=thick, top=left_cell.border.top,
            bottom=left_cell.border.bottom, right=left_cell.border.right
        )
        # Right column
        right_cell = ws.cell(row=row, column=max_col)
        right_cell.border = Border(
            right=thick, top=right_cell.border.top,
            bottom=right_cell.border.bottom, left=right_cell.border.left
        )

    # Thick border around first row (headers)
    for col in range(1, max_col + 1):
        cell = ws.cell(row=1, column=col)
        cell.border = Border(
            top=thick,
            bottom=thick,
            left=cell.border.left,
            right=cell.border.right
        )

    # Thick border around column A
    for row in range(1, max_row + 1):
        cell = ws.cell(row=row, column=1)
        cell.border = Border(
            left=thick,
            right=thick,
            top=cell.border.top,
            bottom=cell.border.bottom
        )

    # Add footer text: "Current Statistics as of Month, Year"
    footer_row = max_row + 2
    current_text = f"Current Statistics as of {datetime.now():%B} 1, {datetime.now():%Y}"
    ws.merge_cells(start_row=footer_row, start_column=1, end_row=footer_row, end_column=max_col)
    cell = ws.cell(row=footer_row, column=1)
    cell.value = current_text
    cell.font = Font(bold=True, size=16)
    cell.alignment = Alignment(horizontal='center', vertical='center')

    wb.save(file_path)
    print(f"Formatted Excel file saved to: {file_path}")

# ---- PROCESS FILE ----

def process_file(file_path):
    convert_file_to_UTF(file_path)
    df = pd.read_csv(file_path, low_memory=False)
    df = df.dropna(axis=1, how='all')

    # Convert Login date and Joined On date to datetime
    df['Last Login'] = pd.to_datetime(df['Last Login'], format="%m/%d/%Y")
    df['Joined On'] = pd.to_datetime(df['Joined On'], format="%m/%d/%Y")
    
    # Get current month and year
    today = datetime.today()
    current_month = today.month
    current_year = today.year

    # Determine previous month and correct year
    if current_month == 1:
        prev_month = 12
        prev_year = current_year - 1
    else:
        prev_month = current_month - 1
        prev_year = current_year
    
    df.loc[:, 'Active Classmates'] = np.where(
            (df['Member Type'] == 'Classmates') &
            (df['Joined On'].isna()== False), 'Active Classmate', pd.NA)

    df.loc[:, 'Total Teachers'] = np.where(
            (df['Member Type'] == 'Teachers') |
            (df['Member Type'] == 'Deceased Teachers'), 'Teacher', pd.NA)

    df.loc[:, 'Total Classmates'] = np.where(
            (df['Member Type'] == 'Classmates') |
            (df['Member Type'] == 'Deceased Classmates'), 'Classmate', pd.NA)

    df.loc[:, 'In Memory'] = np.where(
            (df['Member Type'] == 'Deceased Classmates') |
            (df['Member Type'] == 'Deceased Teachers'), 'In Memory', pd.NA)

    df.loc[:, 'In Memory Classmates'] = np.where(
            df['Member Type'] == 'Deceased Classmates', 'In Memory', pd.NA)

    df.loc[:, 'In Memory Teachers'] = np.where(
            df['Member Type'] == 'Deceased Teachers', 'In Memory', pd.NA)

    df.loc[:, 'Site Visitors (Last Month)'] = np.where(
            (df['Last Login'].dt.month == prev_month) &
            (df['Last Login'].dt.year == prev_year), 'Current', pd.NA)
    
    df.loc[:, 'New Sign-Ups (Last Month)'] = np.where(
            (df['Joined On'].dt.month == prev_month) &
            (df['Joined On'].dt.year == prev_year), 'Current', pd.NA)
    
    stats = df.groupby('School')[[
        'Total Classmates', 'Active Classmates', 'In Memory Classmates',
        'Total Teachers', 'In Memory Teachers', 'Site Visitors (Last Month)', 'New Sign-Ups (Last Month)'
    ]].count()

    stats.loc['TOTAL'] = stats.sum()

    return stats.reset_index()
    
# ---- MAIN ----

def main():
    root = tk.Tk()
    root.withdraw()

    file_path = select_file("Select the Consolidated CSV file")
    if not file_path:
        messagebox.showinfo("Cancelled", "No file selected.")
        return

    stats_df = process_file(file_path)
    save_to_excel(stats_df)

if __name__ == "__main__":
    main()


# In[ ]:





# In[ ]:




